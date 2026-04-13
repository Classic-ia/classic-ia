"""
importador_producao.py — Classic RH & SST
Importa dados de produção: Atak ERP + Planilhas operacionais

Uso:
  python importador_producao.py --atak "caminho/arquivo.xlsx"
  python importador_producao.py --planilha "caminho/arquivo.xlsx"
  python importador_producao.py --consolidar
"""

import sys, os, json, hashlib, requests, datetime, time, re, zipfile
import xml.etree.ElementTree as ET

SB_URL = "https://muiqmtnfvyffborgiwdw.supabase.co"
SB_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im11aXFtdG5mdnlmZmJvcmdpd2R3Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzM4NzQ2MTAsImV4cCI6MjA4OTQ1MDYxMH0.ReerGoDWcLrxJ1WIIKflN4Sc-st4zR1b6yWagCaPq_A"
HEADERS = {"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}", "Content-Type": "application/json"}

def rpc(name, params):
    r = requests.post(f"{SB_URL}/rest/v1/rpc/{name}", headers=HEADERS, json=params)
    return r.json() if r.ok else {"error": r.text[:200]}


def importar_atak(arquivo):
    """Importa produção do Atak ERP via XML (porque openpyxl falha com formato BR)"""
    print(f"[ATAK] Importando: {arquivo}")

    with zipfile.ZipFile(arquivo) as z:
        # Ler strings compartilhadas
        with z.open("xl/sharedStrings.xml") as f:
            tree = ET.parse(f)
            ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            strings = []
            for si in tree.findall(".//s:si", ns):
                text = "".join(t.text or "" for t in si.findall(".//s:t", ns))
                strings.append(text)

        # Ler sheet1
        with z.open("xl/worksheets/Sheet1.xml") as f:
            tree = ET.parse(f)
            rows = tree.findall(".//s:row", ns)

    print(f"  Strings: {len(strings)}, Rows: {len(rows)}")

    registros = []
    data_atual = None

    for row in rows:
        cells = []
        for cell in row.findall("s:c", ns):
            t = cell.get("t")
            val_elem = cell.find("s:v", ns)
            if val_elem is not None:
                if t == "s":
                    idx = int(val_elem.text)
                    cells.append(strings[idx] if idx < len(strings) else "")
                else:
                    cells.append(val_elem.text)
            else:
                cells.append("")

        if not cells:
            continue

        # Detectar data (formato DD/MM/YYYY)
        if len(cells) == 1 and re.match(r"^\d{2}/\d{2}/\d{4}$", cells[0]):
            try:
                data_atual = datetime.datetime.strptime(cells[0], "%d/%m/%Y").strftime("%Y-%m-%d")
            except:
                pass
            continue

        # Ignorar headers e totais
        if not data_atual:
            continue
        if cells[0].startswith("Total") or cells[0] == "Produto" or not cells[0]:
            continue

        # Parsear linha de produção
        produto = cells[0]
        try:
            qtd_kg = float(str(cells[4]).replace(",", ".")) if len(cells) > 4 and cells[4] else 0
            qtd_pcs = int(float(str(cells[8]).replace(",", "."))) if len(cells) > 8 and cells[8] else 0
            unidade = cells[5] if len(cells) > 5 and cells[5] else "KG"
            prod_hora = float(str(cells[9]).replace(",", ".")) if len(cells) > 9 and cells[9] else 0

            # Extrair código do produto
            parts = produto.split("-", 1)
            codigo = parts[0].strip() if parts else ""
            nome = parts[1].strip() if len(parts) > 1 else produto

            registros.append({
                "data": data_atual,
                "produto_codigo": codigo,
                "produto_nome": nome,
                "quantidade_kg": round(qtd_kg, 2),
                "quantidade_pcs": qtd_pcs,
                "unidade": unidade,
                "producao_hora": round(prod_hora, 2),
            })
        except (ValueError, IndexError):
            continue

    print(f"  Registros extraidos: {len(registros)}")
    print(f"  Datas: {len(set(r['data'] for r in registros))}")

    # Enviar em lotes
    total_ins = 0
    for i in range(0, len(registros), 100):
        lote = registros[i:i+100]
        result = rpc("importar_producao_atak", {"p_registros": lote})
        if isinstance(result, list) and result:
            total_ins += result[0].get("inseridos", 0)
            print(f"  Lote {i//100+1}: {result[0]}")
        else:
            print(f"  Lote {i//100+1}: {result}")
        time.sleep(0.3)

    print(f"[DONE] Atak: {total_ins} inseridos")
    return total_ins


def importar_planilha(arquivo):
    """Importa produção por funcionário da planilha operacional"""
    import openpyxl
    print(f"[PLANILHA] Importando: {os.path.basename(arquivo)}")

    wb = openpyxl.load_workbook(arquivo, data_only=True)
    registros = []

    # Extrair referência da semana do nome do arquivo
    semana_ref = os.path.basename(arquivo).replace(".xlsx", "")

    for sheet_name in wb.sheetnames:
        if sheet_name.upper() in ("GERAL",):
            continue  # pular aba geral (totais)

        ws = wb[sheet_name]
        setor = sheet_name.strip()

        # Encontrar datas na linha de cabeçalho (row 1 ou row 3)
        datas = []
        for r in range(1, min(5, ws.max_row + 1)):
            for c in range(2, min(25, ws.max_column + 1)):
                val = ws.cell(r, c).value
                if isinstance(val, datetime.datetime):
                    datas.append((c, val.strftime("%Y-%m-%d")))

        if not datas:
            continue

        # Ler dados dos funcionários (a partir da linha com nomes)
        for r in range(3, ws.max_row + 1):
            nome_cell = ws.cell(r, 1).value
            if not nome_cell or str(nome_cell).startswith("TOTAL") or str(nome_cell).startswith("LEGENDA"):
                continue

            nome_raw = str(nome_cell).strip()
            if not nome_raw or nome_raw == "#DIV/0!" or nome_raw == " ":
                continue

            # Extrair nome e matrícula (ex: "CELIA 86023")
            match = re.match(r"^(.+?)\s+(\d{4,6})$", nome_raw)
            if match:
                nome = match.group(1).strip().upper()
                matricula = match.group(2)
            else:
                nome = nome_raw.upper()
                matricula = None

            # Para cada data, pegar quantidade (coluna da data)
            for col_idx, data_str in datas:
                qtd_val = ws.cell(r, col_idx).value
                if qtd_val is None or str(qtd_val) == "#DIV/0!":
                    continue
                try:
                    qtd = float(qtd_val)
                    if qtd <= 0:
                        continue
                except (ValueError, TypeError):
                    continue

                # Pegar média e horas (colunas seguintes)
                media = 0
                horas = 0
                try:
                    media_val = ws.cell(r, col_idx + 1).value
                    if media_val and str(media_val) != "#DIV/0!":
                        media = float(media_val)
                except:
                    pass
                try:
                    horas_val = ws.cell(r, col_idx + 2).value
                    if horas_val and str(horas_val) != "#DIV/0!":
                        horas = float(horas_val)
                except:
                    pass

                registros.append({
                    "data": data_str,
                    "nome": nome,
                    "matricula": matricula,
                    "setor": setor,
                    "quantidade": round(qtd, 2),
                    "media_hora": round(media, 2),
                    "horas": round(horas, 2),
                    "semana_ref": semana_ref,
                })

    print(f"  Registros extraidos: {len(registros)}")
    print(f"  Setores: {set(r['setor'] for r in registros)}")
    print(f"  Funcionarios: {len(set(r['nome'] for r in registros))}")

    # Enviar em lotes
    total_ins = 0
    total_vinc = 0
    for i in range(0, len(registros), 50):
        lote = registros[i:i+50]
        result = rpc("importar_producao_rateada", {"p_registros": lote})
        if isinstance(result, list) and result:
            total_ins += result[0].get("inseridos", 0)
            total_vinc += result[0].get("vinculados", 0)
            print(f"  Lote {i//50+1}: {result[0]}")
        else:
            print(f"  Lote {i//50+1}: {result}")
        time.sleep(0.3)

    print(f"[DONE] Planilha: {total_ins} inseridos, {total_vinc} vinculados")
    return total_ins


def consolidar():
    """Executa consolidação (reconciliação Atak vs planilha)"""
    print("[CONSOLIDAR] Executando...")
    result = rpc("fn_consolidar_producao", {"p_data": None})
    print(f"  Resultado: {result}")


if __name__ == "__main__":
    if "--atak" in sys.argv:
        idx = sys.argv.index("--atak")
        if idx + 1 < len(sys.argv):
            importar_atak(sys.argv[idx + 1])
    elif "--planilha" in sys.argv:
        idx = sys.argv.index("--planilha")
        if idx + 1 < len(sys.argv):
            importar_planilha(sys.argv[idx + 1])
    elif "--consolidar" in sys.argv:
        consolidar()
    else:
        print("Uso:")
        print('  python importador_producao.py --atak "arquivo.xlsx"')
        print('  python importador_producao.py --planilha "arquivo.xlsx"')
        print("  python importador_producao.py --consolidar")
