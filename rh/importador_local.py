"""
importador_local.py — Classic RH & SST
Importa dados de planilhas Excel da pasta local para o Supabase.

Uso:
  python importador_local.py                # importar tudo
  python importador_local.py --atestados    # apenas atestados
  python importador_local.py --dry-run      # simular sem enviar

Requer: pip install openpyxl requests
"""

import os
import sys
import json
import hashlib
import requests
import datetime
from pathlib import Path

# ═══ CONFIG ═══
SB_URL = "https://muiqmtnfvyffborgiwdw.supabase.co"
SB_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im11aXFtdG5mdnlmZmJvcmdpd2R3Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzM4NzQ2MTAsImV4cCI6MjA4OTQ1MDYxMH0.ReerGoDWcLrxJ1WIIKflN4Sc-st4zR1b6yWagCaPq_A"

BASE_PATH = Path(r"C:\Users\janai\OneDrive\RH e SEGURANÇA DO TRABALHO CLASSIC\RH_E_SEGURANCA_DO_TRABALHO_CLASSIC")

HEADERS = {
    "apikey": SB_KEY,
    "Authorization": f"Bearer {SB_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=minimal",
}

DRY_RUN = "--dry-run" in sys.argv

# ═══ HELPERS ═══
def file_hash(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()

def sb_post(table, data):
    if DRY_RUN:
        print(f"  [DRY] POST {table}: {json.dumps(data, ensure_ascii=False)[:100]}")
        return True
    r = requests.post(f"{SB_URL}/rest/v1/{table}", headers=HEADERS, json=data)
    return r.status_code < 300

def sb_get(table, query=""):
    r = requests.get(f"{SB_URL}/rest/v1/{table}?{query}", headers=HEADERS)
    return r.json() if r.ok else []

def sb_upsert(table, data, on_conflict="id"):
    if DRY_RUN:
        print(f"  [DRY] UPSERT {table}: {len(data)} registros")
        return True
    h = {**HEADERS, "Prefer": "return=minimal,resolution=merge-duplicates"}
    r = requests.post(f"{SB_URL}/rest/v1/{table}", headers=h, json=data)
    return r.status_code < 300

def log_file(caminho, nome, tipo, categoria, status, registros=0, erro=None, hash_arq=None):
    sb_post("file_ingest_log", {
        "caminho": str(caminho),
        "nome_arquivo": nome,
        "tipo_arquivo": tipo,
        "categoria": categoria,
        "status": status,
        "registros_extraidos": registros,
        "erro_detalhe": erro,
        "hash_arquivo": hash_arq,
        "processado_em": datetime.datetime.now().isoformat(),
    })

def ja_processado(hash_arq):
    """Verifica se o arquivo já foi processado pelo hash"""
    r = requests.get(
        f"{SB_URL}/rest/v1/file_ingest_log?hash_arquivo=eq.{hash_arq}&status=eq.processado&select=id",
        headers=HEADERS
    )
    return r.ok and len(r.json()) > 0

def limpar_cpf(cpf):
    if not cpf:
        return None
    return "".join(c for c in str(cpf) if c.isdigit()).zfill(11)

def parse_data(val):
    if not val:
        return None
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"]:
        try:
            return datetime.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


# ═══ IMPORTADOR: ATESTADOS ═══
def importar_atestados():
    """Importa planilha de controle de atestados"""
    import openpyxl

    arquivo = BASE_PATH / "RELATORIO ATESTADO" / "JANEIRO 2025 A JUNHO 2025" / "Controle Atestado.xlsx"
    if not arquivo.exists():
        print(f"[ERRO] Arquivo nao encontrado: {arquivo}")
        return

    h = file_hash(arquivo)
    if ja_processado(h):
        print(f"[SKIP] Atestados ja processado (hash: {h[:12]})")
        return

    print(f"[INIT] Importando atestados: {arquivo.name}")
    wb = openpyxl.load_workbook(arquivo, data_only=True)
    ws = wb.active

    total = 0
    erros = 0
    inseridos = 0
    nao_encontrados = 0

    # Buscar funcionarios do core para match por CPF (via RPC SECURITY DEFINER)
    r = requests.post(f"{SB_URL}/rest/v1/rpc/rh_cpf_map", headers=HEADERS, json={})
    funcionarios = r.json() if r.ok else []
    cpf_map = {f["cpf"]: f["id"] for f in funcionarios if f.get("cpf")}
    print(f"  Funcionarios no core: {len(cpf_map)}")

    batch = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        total += 1

        nome = str(row[1] or "").strip()
        cpf = limpar_cpf(row[2])
        motivo = str(row[3] or "")
        tipo = str(row[4] or "")
        data_de = parse_data(row[5])
        data_ate = parse_data(row[6])
        justificativa = str(row[7] or "")
        cid = str(row[8] or "")
        dias = int(row[9]) if row[9] and str(row[9]).isdigit() else 1
        medico = str(row[10] or "")
        hospital = str(row[11] or "")

        func_id = cpf_map.get(cpf)
        if not func_id:
            nao_encontrados += 1
            continue

        payload = {
            "funcionario_id": func_id, "nome": nome, "cpf": cpf,
            "motivo": motivo, "tipo": tipo, "data_inicio": data_de,
            "data_fim": data_ate, "justificativa": justificativa,
            "cid": cid if cid and "apresentou" not in cid.lower() else None,
            "dias": dias, "medico": medico, "hospital": hospital,
        }

        batch.append({
            "id_origem": f"atestado_{cpf}_{data_de}",
            "payload_json": payload,
            "hash_registro": hashlib.sha256(json.dumps(payload, sort_keys=True).encode()).hexdigest(),
        })

    # Enviar via RPC (SECURITY DEFINER, bypassa RLS)
    print(f"  Enviando {len(batch)} registros via RPC...")
    import time
    for i in range(0, len(batch), 50):
        lote = batch[i:i+50]
        try:
            r = requests.post(f"{SB_URL}/rest/v1/rpc/file_ingest_atestados",
                headers=HEADERS, json={"p_registros": lote})
            if r.ok:
                result = r.json()
                if isinstance(result, list) and result:
                    inseridos += result[0].get("inseridos", 0)
                else:
                    inseridos += len(lote)
            else:
                print(f"  [ERRO] Lote {i//50+1}: HTTP {r.status_code} {r.text[:100]}")
                erros += len(lote)
        except Exception as e:
            print(f"  [ERRO] Lote {i//50+1}: {e}")
            erros += len(lote)
        time.sleep(0.3)

    print(f"[DONE] Atestados: {total} lidos, {inseridos} staging, {nao_encontrados} sem match CPF, {erros} erros")
    log_file(arquivo, arquivo.name, "xlsx", "atestado", "processado", inseridos, hash_arq=h)


# ═══ IMPORTADOR: EXTINTORES ═══
def importar_extintores():
    """Importa planilha de extintores"""
    import openpyxl

    arquivo = BASE_PATH / "SEGURANÇA DO TRABALHO" / "Extintores_Organizados - TROCADO 2025.xlsx"
    if not arquivo.exists():
        print(f"[ERRO] Arquivo nao encontrado: {arquivo}")
        return

    h = file_hash(arquivo)
    if ja_processado(h):
        print(f"[SKIP] Extintores ja processado (hash: {h[:12]})")
        return

    print(f"[INIT] Importando extintores: {arquivo.name}")
    wb = openpyxl.load_workbook(arquivo, data_only=True)
    ws = wb.active

    print(f"  Sheets: {wb.sheetnames}")
    print(f"  Rows: {ws.max_row}, Cols: {ws.max_column}")

    # Mostrar headers para entender estrutura
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        print(f"  {[str(c)[:30] if c else '' for c in row]}")

    log_file(arquivo, arquivo.name, "xlsx", "outro", "processado", 0, hash_arq=h)
    print(f"[DONE] Extintores: estrutura mapeada")


# ═══ INVENTÁRIO GERAL ═══
def inventario():
    """Lista todos os Excel na pasta e classifica"""
    print(f"\n{'='*60}")
    print(f"INVENTARIO DE ARQUIVOS EXCEL")
    print(f"{'='*60}\n")

    excels = list(BASE_PATH.rglob("*.xlsx")) + list(BASE_PATH.rglob("*.xls"))
    print(f"Total: {len(excels)} arquivos Excel\n")

    categorias = {
        "atestado": [], "folha": [], "aso": [], "treinamento": [],
        "epi": [], "admissao": [], "desligamento": [], "outro": []
    }

    for f in sorted(excels):
        rel = str(f.relative_to(BASE_PATH)).lower()
        nome = f.name.lower()

        if "atestado" in rel or "atestado" in nome:
            categorias["atestado"].append(f)
        elif "folha" in rel or "pgto" in nome or "pagamento" in rel:
            categorias["folha"].append(f)
        elif "exame" in rel or "aso" in nome or "periodico" in rel:
            categorias["aso"].append(f)
        elif "treinamento" in rel or "curso" in rel or "nr" in nome:
            categorias["treinamento"].append(f)
        elif "epi" in rel or "epi" in nome:
            categorias["epi"].append(f)
        elif "admiss" in rel or "admissional" in nome:
            categorias["admissao"].append(f)
        elif "deslig" in rel:
            categorias["desligamento"].append(f)
        else:
            categorias["outro"].append(f)

    for cat, files in categorias.items():
        if files:
            print(f"\n--- {cat.upper()} ({len(files)} arquivos) ---")
            for f in files[:5]:
                size = f.stat().st_size // 1024
                print(f"  [{size}KB] {f.relative_to(BASE_PATH)}")
            if len(files) > 5:
                print(f"  ... +{len(files)-5} mais")


# ═══ MAIN ═══
if __name__ == "__main__":
    print(f"Classic RH & SST — Importador Local")
    print(f"Base: {BASE_PATH}")
    print(f"Supabase: {SB_URL}")
    print(f"Dry run: {DRY_RUN}")
    print()

    if "--inventario" in sys.argv:
        inventario()
    elif "--atestados" in sys.argv:
        importar_atestados()
    elif "--extintores" in sys.argv:
        importar_extintores()
    elif "--all" in sys.argv:
        importar_atestados()
        importar_extintores()
    else:
        print("Uso:")
        print("  python importador_local.py --inventario    # listar arquivos")
        print("  python importador_local.py --atestados     # importar atestados")
        print("  python importador_local.py --all           # importar tudo")
        print("  python importador_local.py --dry-run --all # simular")
